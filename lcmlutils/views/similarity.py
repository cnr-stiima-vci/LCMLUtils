import re
import os
import sys
import copy
import lxml
from lxml import etree
from lxml import objectify
import pprint
import operator
import requests
from operator import itemgetter
import itertools
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import (get_column_letter)
from difflib import SequenceMatcher as SM
from pdb import set_trace as _breakpoint

from django.http import HttpResponse, JsonResponse 
from django.views.decorators.csrf import csrf_exempt
 
from lcmlutils.settings import DATA_PATH, BFIS_SERVICE_PORT
from lcmlutils.models import LCCS3Class, LCCS3Legend
from lcmlutils.views.transcoding import transcode_lccs3_classes, namespaces
from lcmlutils.views.multiverse import generate_valid_permutations, get_presence_type_groups

biotic_sim_dict = {}
extensiveness_dict = {}
 
def load_similarity_table_and_names(fn):
    dd = {}
    names = []
    #print(fn)
    wb = load_workbook(filename = fn)
    ws = wb.active
    try:
        for r_idx in range(ws.min_row, ws.max_row+1):
            rawname = ws['a{0}'.format(r_idx)].value
            name = re.split(' |\(',rawname)[0]
            dd[name] = {}
            names.append(name)
        for r_idx in range(len(names)):
            for c_idx in range(len(names)):
                # cell_pos = chr(ord('a')+c_idx+1)+str(r_idx+1)
                cell_pos = get_column_letter(c_idx+2)+str(r_idx+1)
                val = ws[cell_pos].value
                dd[names[r_idx]][names[c_idx]] = val
    except Exception:
        #print(e)
        #_breakpoint()
        pass
    return dd, names
 
 
def collect_table_files_from_folder(files_path):
    f = []
    for (dirpath, dirnames, filenames) in os.walk(files_path):
        cfns = [os.path.join(dirpath, lfn) for lfn in filenames]
        #print(cfns)
        f.extend(cfns)
        break
    return f
 
 
def compute_phase1_score(ref_class, query_class, names, dd, ed):
    ref_class_cnt = len(ref_class)
    query_class_cnt = len(query_class)
    cnt1 = 0
    final_score = 0
    phi_scores = []
    eps_scores = []
    default_value = 1
    ref_class_names = [qe.get('element_type') for qe in ref_class]
    for ec2 in query_class:
        #ref_elem = m2n(ec2, names)
        query_elem = ec2.get('element_type')
        #scores = [dd[ref_elem][m2n(ec1,names)]for ec1 in ref_class] #symmetric matrix dd[a][b]==dd[b][a]
        scores = [dd.get(query_elem,{}).get(qe, default_value) for qe in ref_class_names]
        try:
            phi_score = max(scores)
            phi_scores.append(phi_score)
            phi_index = scores.index(phi_score)
            del ref_class_names[phi_index]
            #eps_score = phi_score * weights_mapping_dict[ref_class_cnt][phi_index]
            #eps_scores.append(eps_score)
            cnt1+=1
        except Exception:
            pass
    print('ref class: {0}'.format(ref_class))
    print('query class: {0}'.format(query_class))
    print('phi scores: {0}'.format(phi_scores))
    phi_score = mean(phi_scores)
    print('phi score: {0}'.format(phi_score))
    #extensiveness_weight = ed[query_class_cnt][ref_class_cnt]
    extensiveness_weight = ed.get(query_class_cnt,{}).get(ref_class_cnt,0)
    psi_score = phi_score * extensiveness_weight
    final_score = psi_score
    return final_score


def count_props_with_name_and_value_in_transcoded_class(transcoded_class, prop_name, value):
    return sum(
        elem['properties'][propname].get('attributes').get('value')==value \
            for elem in transcoded_class \
                for propname in elem.get('properties').keys() \
                    if propname==prop_name)    

def count_basic_elements_with_name_and_value_including_additional_propname_in_transcoded_class(transcoded_class, prop_name, value, prop2_name):
    return sum(
        elem['properties'][propname].get('attributes').get('value')==value \
            for elem in transcoded_class \
                for propname in elem.get('properties').keys() \
                    if propname==prop_name and prop2_name in elem.get('properties').keys())    

def collect_values_for_props_with_name_in_transcoded_class(transcoded_class, prop_name, attr_name):
    return [
        elem['properties'][propname].get('attributes').get(attr_name) \
            for elem in transcoded_class \
                for propname in elem.get('properties').keys() \
                    if propname==prop_name]

def collect_values_for_props_with_name_having_additional_propname_in_transcoded_class(transcoded_class, prop_name, attr_name, prop2_name):
    return [
        elem['properties'][propname].get('attributes').get(attr_name) \
            for elem in transcoded_class \
                for propname in elem.get('properties').keys() \
                    if propname==prop_name and prop2_name in elem.get('properties').keys()]


def compute_property_extensiveness(transcoded_class, prop_name):
    extensiveness_length = len(transcoded_class)
    excl_elems_found = count_basic_elements_with_name_and_value_including_additional_propname_in_transcoded_class(transcoded_class, 'presence_type', 'Exclusive', prop_name)    
    tempseq_elems_found = count_basic_elements_with_name_and_value_including_additional_propname_in_transcoded_class(transcoded_class, 'presence_type', 'Temporal Sequence Depending', prop_name)
    mandatory_elems_found = count_basic_elements_with_name_and_value_including_additional_propname_in_transcoded_class(transcoded_class, 'presence_type', 'Mandatory', prop_name)
    opt_elems_found = count_basic_elements_with_name_and_value_including_additional_propname_in_transcoded_class(transcoded_class, 'presence_type', 'Optional', prop_name)
    print("exclusive: {0}; mandatory: {1}; temporal: {2}; optional {3}".format(excl_elems_found, mandatory_elems_found, tempseq_elems_found, opt_elems_found))
    if excl_elems_found>0:
        extensiveness_length = 1
    else:
        extensiveness_length = mandatory_elems_found
        if tempseq_elems_found:
            tempseq_types = collect_values_for_props_with_name_having_additional_propname_in_transcoded_class(transcoded_class, 'sequential_temporal_relationship', 'type',prop_name)
            tempseq_sy_found = sum([v == 'Sequential Same Year' for v in tempseq_types])
            print("tempseq sy: {0}".format(tempseq_sy_found))
            extensiveness_length += tempseq_sy_found if tempseq_sy_found>0 else 1
    print("extensiveness_length: {0}".format(extensiveness_length))
    return extensiveness_length


def compute_extensiveness(transcoded_class):
    extensiveness_length = len(list(transcoded_class))
    excl_elems_found = count_props_with_name_and_value_in_transcoded_class(transcoded_class, 'presence_type', 'Exclusive')    
    tempseq_elems_found = count_props_with_name_and_value_in_transcoded_class(transcoded_class, 'presence_type', 'Temporal Sequence Depending')
    mandatory_elems_found = count_props_with_name_and_value_in_transcoded_class(transcoded_class, 'presence_type', 'Mandatory')
    opt_elems_found = count_props_with_name_and_value_in_transcoded_class(transcoded_class, 'presence_type', 'Optional')
    print("exclusive: {0}; mandatory: {1}; temporal: {2}; optional {3}".format(excl_elems_found, mandatory_elems_found, tempseq_elems_found, opt_elems_found))
    if excl_elems_found>0:
        extensiveness_length = 1 + mandatory_elems_found #+ mandatory_elems_found
    else:
        extensiveness_length = mandatory_elems_found
        if tempseq_elems_found:
            tempseq_types = collect_values_for_props_with_name_in_transcoded_class(transcoded_class, 'sequential_temporal_relationship', 'type',)
            tempseq_sy_found = sum([v == 'Sequential Same Year' for v in tempseq_types])
            print("tempseq sy: {0}".format(tempseq_sy_found))
            extensiveness_length += tempseq_sy_found if tempseq_sy_found>0 else 1
    print("extensiveness_length: {0}".format(extensiveness_length))
    return extensiveness_length

def apply_extensiveness_penalty(starting_score, query_class, ref_class, matching_pairs):
    # apply penalty for mandatory elements in the query class that do not appear in any matching pair
    new_score = starting_score
    mandatory_elems_total = 0
    mandatory_elems_total += count_props_with_name_and_value_in_transcoded_class(query_class, 'presence_type', 'Mandatory')
    mandatory_elems_total += count_props_with_name_and_value_in_transcoded_class(ref_class, 'presence_type', 'Mandatory')
    if mandatory_elems_total:
        mandatory_matched = 0
        for mp in matching_pairs:
            ptval = mp['qe_props']['presence_type']['attributes']['value']
            if ptval == 'Mandatory':
                mandatory_matched+=1
            ptval = mp['re_props']['presence_type']['attributes']['value']
            if ptval == 'Mandatory':
                mandatory_matched+=1
        missing_count = mandatory_elems_total - mandatory_matched
        if missing_count>0:
            perc = missing_count/float(mandatory_elems_total)
            perc = perc*0.5 #maximum penalty is half the computed extensiveness
            new_score = new_score*(1.0 - perc)
    return new_score


def create_class_dict(transcoded_class):
    dd = {}
    for elem in transcoded_class:
        dd[elem.get('element_uuid')] = elem
    return dd

def select_partial_class(transcoded_class_dict, uuids):
    dd = transcoded_class_dict
    partial_class = []
    for uuid in uuids:
        if uuid in dd.keys():
            partial_class.append(dd[uuid])
    return partial_class

def compute_phase1_with_multiverse(ref_class, query_class, names, dd, ed):
    ref_class_cnt = compute_extensiveness(ref_class)
    cnt1 = 0
    final_score = 0
    default_value = 1
    accepted_pt = ['Mandatory','Exclusive'] #, 'Temporal Sequence Depending'
    orig_ref_class_names = [rec.get('element_type') for rec in ref_class]
    orig_ref_class_uuids = [rec.get('element_uuid') for rec in ref_class]
    orig_ref_class_props = [rec.get('properties') for rec in ref_class]
    groups = get_presence_type_groups(query_class)
    mg = groups['Mandatory']
    og = groups['Optional']
    eg = groups['Exclusive']
    permutation_scores = []
    tqcd = create_class_dict(query_class)
    permutations = generate_valid_permutations(mg, og, eg, greedy_mode = True)
    
    for permutation in permutations:
        partial_qc = select_partial_class(tqcd, permutation)
        query_class_cnt = compute_extensiveness(partial_qc)
        orig_query_class_names = [qe.get('element_type') for qe in partial_qc]
        orig_query_class_uuids = [qe.get('element_uuid') for qe in partial_qc]
        
        phi_scores = []
        eps_scores = []
        matching_pairs = []
        query_class_names = copy.copy(orig_query_class_names)
        query_class_uuids = copy.copy(orig_query_class_uuids)
        query_class_props = [qe.get('properties') for qe in partial_qc]
        ref_class_names = copy.copy(orig_ref_class_names)
        ref_class_uuids = copy.copy(orig_ref_class_uuids)
        ref_class_props = copy.copy(orig_ref_class_props)
        ref_class_positions = list(range(len(ref_class_names)))

        qidx = 0
        for query_elem in query_class_names:
            pprint.pprint(query_elem)
            scores = [dd.get(query_elem,{}).get(rcn, default_value) for rcn in ref_class_names]
            pprint.pprint('vs {0}'.format(ref_class_names))
            pprint.pprint(scores)
            if len(scores)>0:
                try:
                    max_score = max(scores)
                    #phi_score = sum(scores)/float(len(scores))
                    phi_score = max_score
                    phi_index = scores.index(max_score)
                    ref_class_pos = ref_class_positions[phi_index]
                    portioning_rc = ref_class[ref_class_pos].get('properties').get('portioning',{'attributes':{'min':100}}).get('attributes').get('min')
                    phi_scores.append(phi_score * portioning_rc/100.0)
                    qe_props = list(query_class_props)[qidx]
                    matching_pairs.append(
                        {
                            'qe_uuid': list(query_class_uuids)[qidx], 
                            'qe_type': query_elem,
                            'qe_props':qe_props, 
                            're_uuid': ref_class_uuids[phi_index],
                            're_type': ref_class_names[phi_index],
                            're_props':ref_class_props[phi_index]
                        })
                    del ref_class_names[phi_index]
                    del ref_class_uuids[phi_index]
                    del ref_class_positions[phi_index]
                    #eps_score = phi_score * weights_mapping_dict[ref_class_cnt][phi_index]
                    #eps_scores.append(eps_score)
                    cnt1+=1
                    is_seq_sy_elem = ref_class[ref_class_pos].get('properties').get('sequential_temporal_relationship',{'attributes':{'type':None}}).get('attributes').get('type')=='Sequential Same Year'
                    if is_seq_sy_elem:
                        query_class_cnt = len(phi_scores)
                        break
                except:
                    err = sys.exc_info()[0]
                    pass
            else:
                pprint.pprint("no matches")
            qidx+=1
            print(phi_scores)
        
        phi_score = mean(phi_scores) #sum(phi_scores) 
        print('phi score: {0}'.format(phi_score))
        #extensiveness_weight = ed[query_class_cnt][ref_class_cnt]
        print("ed:")
        print(ed)
        extensiveness_weight = ed.get(query_class_cnt,{}).get(ref_class_cnt,1) or 1
        print('extensiveness {0} vs {1} = {2}'.format(ref_class_cnt, query_class_cnt, extensiveness_weight))
        extensiveness_final = apply_extensiveness_penalty(extensiveness_weight, query_class, ref_class, matching_pairs)
        #psi_score = phi_score * extensiveness_weight
        psi_score = phi_score * extensiveness_final
        final_score = psi_score
        print('final_score: {0}'.format(final_score))
        permutation_scores.append({
            'query_class_names': query_class_names,
            'permutation': permutation,
            'score': final_score,
            'matching_pairs': matching_pairs,
            'partial_qc':partial_qc})
    return permutation_scores

def compute_phase1_ml(ref_class, query_class, names, dd, ed):
    ref_class_cnt = compute_extensiveness(ref_class)
    query_class_cnt = compute_extensiveness(query_class)
    cnt1 = 0
    final_score = 0
    default_value = 1
    accepted_pt = ['Mandatory','Exclusive','Optional'] #, 'Temporal Sequence Depending'
    orig_ref_class_names = [rec.get('element_type') for rec in ref_class]
    orig_ref_class_uuids = [rec.get('element_uuid') for rec in ref_class]
    orig_ref_class_props = [rec.get('properties') for rec in ref_class]
    orig_query_class_names = [qe.get('element_type') for qe in query_class if qe['properties']['presence_type']['attributes']['value'] in accepted_pt]
    orig_query_class_uuids = [qe.get('element_uuid') for qe in query_class if qe['properties']['presence_type']['attributes']['value'] in accepted_pt]
    permutations = list(itertools.permutations(range(len(orig_query_class_names)), len(orig_query_class_names)))
    pprint.pprint(permutations)
    #_breakpoint()
    permutation_scores = []
    for permutation in permutations:
        pprint.pprint(permutation)
        phi_scores = []
        eps_scores = []
        matching_pairs = []
        query_class_names = map(lambda i: orig_query_class_names[i], permutation)
        query_class_uuids = map(lambda i: orig_query_class_uuids[i], permutation)
        query_class_props = map(lambda i: query_class[i]['properties'], permutation)
        ref_class_names = copy.copy(orig_ref_class_names)
        ref_class_uuids = copy.copy(orig_ref_class_uuids)
        ref_class_props = copy.copy(orig_ref_class_props)
        ref_class_positions = list(range(len(ref_class_names)))
        #print('permutation: {0}'.format(query_class_names))
        qidx = 0
        for query_elem in query_class_names:
            pprint.pprint(query_elem)
            scores = [dd.get(query_elem,{}).get(rcn, default_value) for rcn in ref_class_names]
            pprint.pprint('vs {0}'.format(ref_class_names))
            pprint.pprint(scores)
            if len(scores)>0:
                try:
                    max_score = max(scores)
                    phi_score = sum(scores)/float(len(scores))
                    phi_index = scores.index(max_score)
                    ref_class_pos = ref_class_positions[phi_index]
                    portioning_rc = ref_class[ref_class_pos].get('properties').get('portioning',{'attributes':{'min':100}}).get('attributes').get('min')
                    phi_scores.append(phi_score * portioning_rc/100.0)
                    qe_props = list(query_class_props)[qidx]
                    matching_pairs.append(
                        {
                            'qe_uuid': list(query_class_uuids)[qidx], 
                            'qe_type': query_elem,
                            'qe_props':qe_props, 
                            're_uuid': ref_class_uuids[phi_index],
                            're_type': ref_class_names[phi_index],
                            're_props':ref_class_props[phi_index]
                        })
                    del ref_class_names[phi_index]
                    del ref_class_uuids[phi_index]
                    del ref_class_positions[phi_index]
                    #eps_score = phi_score * weights_mapping_dict[ref_class_cnt][phi_index]
                    #eps_scores.append(eps_score)
                    cnt1+=1
                    is_seq_sy_elem = ref_class[ref_class_pos].get('properties').get('sequential_temporal_relationship',{'attributes':{'type':None}}).get('attributes').get('type')=='Sequential Same Year'
                    if is_seq_sy_elem:
                        query_class_cnt = len(phi_scores)
                        break
                except:
                    err = sys.exc_info()[0]
                    pass
            else:
                pprint.pprint("no matches")
            qidx+=1
        #print('ref class: {0}'.format(ref_class))
        #print('query class: {0}'.format(query_class))
        #print('phi scores: {0}'.format(phi_scores))
        print(phi_scores)
        phi_score = mean(phi_scores) #mean(phi_scores)
        print('phi score: {0}'.format(phi_score))
        #extensiveness_weight = ed[query_class_cnt][ref_class_cnt]
        print("ed:")
        print(ed)
        extensiveness_weight = ed.get(query_class_cnt,{}).get(ref_class_cnt,1) or 1
        print('extensiveness {0} vs {1} = {2}'.format(ref_class_cnt, query_class_cnt, extensiveness_weight))
        #_breakpoint()
        extensiveness_final = apply_extensiveness_penalty(extensiveness_weight, query_class, ref_class, matching_pairs)
        #psi_score = phi_score * extensiveness_weight
        psi_score = phi_score * extensiveness_final
        final_score = psi_score
        print('final_score: {0}'.format(final_score))
        #_breakpoint()
        permutation_scores.append({
            'query_class_names': query_class_names,
            'permutation': permutation,
            'score': final_score,
            'matching_pairs': matching_pairs,
            'partial_qc': query_class})
    return permutation_scores
 
 
def overlap_score(x1,x2,y1,y2):
    max_range_length = float(max(abs(x2-x1),abs(y2-y1)))
    score = 0
    #print('max_range_length: {0}'.format(max_range_length))
    osc = (min(x2,y2) - max(x1,y1))
    if osc < 0:
        osc = 0
    #print('osc: {0}'.format(osc))
    if max_range_length>0:
        score = (osc/max_range_length)*9 + 1
    return(score) 


props_mapping = {
    'cover': {
        'LC_Trees': 'LC_TreeCover',
        'LC_Shrubs': 'LC_ShrubCover',
        'LC_HerbaceousGrowthForms': 'LC_HerbaceousGrowthFormsCover',
        'LC_WoodyGrowthForms': 'LC_WoodyCover',
        'LC_Graminae': 'LC_GraminaeCover',
        'LC_Forbs': 'LC_ForbsCover',
        'LC_Lichen': 'LC_LichenCover',
        'LC_Algae': 'LC_AlgaeCover',
        'LC_GrowthForms': 'LC_GrowthFormsCover'
    },
    'height': {
        'LC_Trees': 'LC_TreeHeight',
        'LC_Shrubs': 'LC_ShrubHeight',
        'LC_HerbaceousGrowthForms': 'LC_HerbaceousGrowthFormsHeight',
        'LC_WoodyGrowthForms': 'LC_WoodyHeight',
        'LC_Graminae': 'LC_GraminaeHeight',
        'LC_Forbs': 'LC_ForbsHeight',
        'LC_Algae': 'LC_AlgaeHeight',
        'LC_GrowthForms': 'LC_GrowthFormsHeight'
    }
}


IGNORED_PROPS_LIST_STEP2 = ['presence_type', 'portioning','sequential_temporal_relationship',]
 
 
def compute_phase2_score(ref_class, query_class, names, dd, ed, phase1_meta={}):
    print("ref class")
    pprint.pprint(ref_class)
    print("query class")
    pprint.pprint(query_class)
    ref_class_cnt = compute_extensiveness(ref_class)
    query_class_cnt = compute_extensiveness(query_class)
    query_props_cnt = 0
    ref_props_cnt = 0
    cnt1 = 0
    final_score = 0
    default_value = 1
    ref_class_props = []
    query_class_props = []
    prop_scores_dd = {}
    matching_pairs = copy.deepcopy(phase1_meta.get('matching_pairs',[]) or [])
    refc_elems = ref_class
    qc_elems = query_class
    for mp in matching_pairs:
        print('matching pair {0}'.format(mp))
        #re = [rce for rce in refc_elems if rce.get('element_uuid')==mp['re_uuid']][0]
        re_props = mp.get('re_props')
        ret = mp.get('re_type')
        pprint.pprint(ret)

        #qe = [qce for qce in qc_elems if qce.get('element_uuid')==mp['qe_uuid']][0]
        #qe_props = qe.get('properties')
        #qet = qe.get('element_type')
        qe_props = mp.get('qe_props')
        qet = mp.get('qe_type')
        
        pprint.pprint(qet)
        #_breakpoint()
        for pn in qe_props.keys():
            print('analyzing property {0}...'.format(pn))
            if pn not in IGNORED_PROPS_LIST_STEP2:
                query_props_cnt += 1
                if pn not in prop_scores_dd.keys():
                    prop_scores_dd[pn] = {"values":[], "extensiveness_score":0}
                if pn in re_props.keys():
                    pprint.pprint(pn)
                    pprint.pprint('matched...')
                    qce = compute_property_extensiveness(query_class, pn)
                    rce = compute_property_extensiveness(ref_class, pn)
                    prop_scores_dd[pn]['extensiveness_score'] = (ed.get(qce,{}).get(rce, default_value) or default_value)
                    #_breakpoint()
                    o_weight = 1.0
                    ref_props_cnt += 1
                    qp = qe_props[pn]
                    rp = re_props[pn]
                    if pn in props_mapping.keys(): # and ret==qet:
                        pmap = props_mapping.get(pn)
                        qpmap = pmap.get(qet, qet) or qet
                        rpmap = pmap.get(ret, ret) or ret
                        o_weight = (dd.get(qpmap,{}).get(rpmap, default_value) or default_value)/10.0
                    else:
                        o_weight = 1.0
                    pprint.pprint('extensiveness-> Query: {0}/ Ref: {1}'.format(qce,rce))
                    pprint.pprint('Property correspondence score: {0}'.format(o_weight))
                    #_breakpoint()
                    if o_weight<0.5 and qet!=ret:
                        # ch 3.2, scenario 1
                        pprint.pprint('ch 3.2, scenario 1 condition: oscore->0.1')
                        o_score=0.1
                    else:
                        if qp.get('ranged_type')==True:
                            o_score = overlap_score(int(qp['attributes']['min']), int(qp['attributes']['max']), 
                                                    int(rp['attributes']['min']), int(rp['attributes']['max']))
                            pprint.pprint('oscore: {0}'.format(o_score))
                        else:
                            o_score = 10 if qp['attributes'].get('value')==rp['attributes'].get('value') else 0
                        o_score = o_score * o_weight   
                    #_breakpoint()                 
                    prop_scores_dd[pn]['values'].append(o_score)
    final_score = None
    pprint.pprint(prop_scores_dd.keys())
    #_breakpoint()
    if len(prop_scores_dd.keys())>0:
        partials = []
        for k in prop_scores_dd.keys():
            entry = prop_scores_dd[k]
            extensiveness_score = entry['extensiveness_score']
            values = entry['values']
            if len(values)>0:
                mean_value = mean(values)
                partials.append(mean_value*extensiveness_score)

        prop_score = mean(partials)
        #pprint.pprint('mean prop score: {0}'.format(prop_score))
        #extensiveness_weight = ed.get(query_props_cnt,{}).get(ref_props_cnt,1) or 1
        #pprint.pprint('extensiveness: {0}'.format(extensiveness_weight))
        #psi_score = prop_score * extensiveness_weight
        #final_score = psi_score
        final_score = prop_score
        print('final score props: {0}'.format(final_score))
    return final_score


def load_extensiveness_dict(fn):
    dd = {}
    wb = load_workbook(filename = fn)
    ws = wb.active
    for r_idx in range(ws.min_row, ws.max_row+1):
        k1 = ws['a{0}'.format(r_idx+1)].value
        k2 = ws['b{0}'.format(r_idx+1)].value
        v  = ws['c{0}'.format(r_idx+1)].value
        if k1 not in dd.keys():
            dd[k1] = {}
        dd[k1][k2] = v
    return dd
 
 
def mean(numbers):
    return float(sum(numbers)) / max(len(numbers), 1)
 
 
def load_correspondence_mtx(corr_type):
    corr_path = os.path.join(DATA_PATH, 'correspondence', corr_type)
    fns = collect_table_files_from_folder(corr_path)
    dd_mtx = {}
    all_names = []
    for fn in fns:
        dd, names = load_similarity_table_and_names(fn)
        dd_mtx.update(dd)
        all_names.extend(names)
    return dd_mtx, all_names
 
 
elements_corr_dict, element_names = load_correspondence_mtx('elements')
props_corr_dict, prop_names = load_correspondence_mtx('properties')
 
el_ext_path = os.path.join(DATA_PATH, 'extensiveness')

ext_fns = collect_table_files_from_folder(el_ext_path)
extensiveness_dict = None
for fn in ext_fns:
    extensiveness_dict = load_extensiveness_dict(fn)


def get_legend_description(doc_legend):
    dd = {'class_codes':[],'names_dict':{}}
    lc_classes = doc_legend.findall('elements/LC_LandCoverClass/map_code', namespaces)
    for lc in lc_classes:
        dd['class_codes'].append(lc.text)
        name = lc.getparent().find('name').text
        dd['names_dict'][lc.text] = name
    return dd

def perform_assessment(wl, qcm, options = {}):
    dd = {}
    similarity_level = options.get('similarity_level') or 'elements+props'
    phase1_logic = options.get('phase1_logic') or 'variants'
    doc = etree.fromstring(wl.xml_text)
    #lc_classes = doc.findall('elements/LC_LandCoverClass/map_code', namespaces)
    #dd['working_classes'] = [lc.text for lc in lc_classes]
    dld = get_legend_description(doc)
    dd['working_classes'] = dld['class_codes']
    dd['working_names_dict'] = dld['names_dict']
    doc = etree.fromstring(qcm.xml_text)
    #lc_classes = doc.findall('elements/LC_LandCoverClass/map_code', namespaces)
    #dd['query_classes'] = [lc.text for lc in lc_classes]
    dld = get_legend_description(doc)
    dd['query_classes'] = dld['class_codes']
    dd['query_names_dict'] = dld['names_dict']
    dd['scores'] = {}
    if wl and qcm:
        query_classes = transcode_lccs3_classes(qcm.xml_text)
        pprint.pprint(query_classes)
        classes = transcode_lccs3_classes(wl.xml_text)
        for qc in query_classes:
            query_class = qc.get('elements')
            #pprint.pprint(classes)
            scores_kv = {}
            for cl in classes:
                ref_class = cl.get('elements')
                if phase1_logic=='variants':
                    permutation_scores = compute_phase1_with_multiverse(ref_class, query_class, element_names, 
                            elements_corr_dict, extensiveness_dict)
                else:
                    #_breakpoint()
                    permutation_scores = compute_phase1_ml(ref_class, query_class, element_names, 
                            elements_corr_dict, extensiveness_dict)
                newlist = sorted(permutation_scores,key=itemgetter('score'), reverse=True)
                score_phase1 = newlist[0]['score']
                matching_pairs = newlist[0]['matching_pairs']
                pprint.pprint('score phase1: {0}'.format(score_phase1))
                #ml_query_class = list(map(lambda i: query_class[i], newlist[0]['permutation']))
                #ml_query_class = copy.copy(query_class[newlist[0]['permutation'][0]])
                ml_query_class = newlist[0]['partial_qc']
                total_score = 0
                if similarity_level=='elements+props':
                    score_phase2 = compute_phase2_score(ref_class, ml_query_class, prop_names, 
                                        props_corr_dict, extensiveness_dict, newlist[0])
                    if score_phase2:
                        pprint.pprint('score phase2: {0}'.format(score_phase2))
                        total_score = score_phase1 * 0.6 + score_phase2 * 0.4
                    else:
                        pprint.pprint('no computation possible for phase2, skipping')
                        total_score = score_phase1
                else:
                    total_score = score_phase1
                #if total_score>100: # FIXME: usually happens with multiple basic elemenets in horizontal patterns
                #    total_score = 100.0
                print('total score: {0}'.format(total_score))
                scores_kv[cl.get('map_code')] = total_score
            dd['scores'][qc.get('map_code')] = scores_kv
    return dd


xml_legend_template_part1 = '''
<LC_Legend xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" id="1" lccs_sort="0" uuid="4996ffb0-1f8a-11e7-aa16-080027c8b774" xsi:noNamespaceSchemaLocation="sim_assessment_test1_query_classes.xsd" xsi:type="LC_Legend">
   <description>test</description>
   <name>test</name>
   <elements>
'''

xml_legend_template_part2 = '''
   </elements>
</LC_Legend>
'''


def invert_payload_order(original_dd):
    dd = {
        'scores':{}, 
        'working_classes':original_dd.get('working_classes'),
        'query_classes':original_dd.get('query_classes'),
        'working_names_dict':original_dd.get('working_names_dict'),
        'query_names_dict':original_dd.get('query_names_dict')
    }
    scores_dd = original_dd.get('scores')
    for k in scores_dd.keys():
        for k2 in scores_dd[k].keys():
            if k2 not in dd['scores'].keys():
                dd['scores'][k2] = {}
            dd['scores'][k2][k] = original_dd['scores'][k][k2]
    return dd

@csrf_exempt
def similarity_assessment(request):
    dd = {}
    if request.method == 'POST':
        params = json.loads(request.body)
        pprint.pprint(params)
        working_legend_name = params.get('working_legend_name')
        lcml_class_name = params.get('lcml_class_name')
        lcml_query_legend_name = params.get('query_legend_name')
        lcml_snippet = params.get('lcml_class')
        advanced_options = params.get('advanced_options') or {}
        wl = None
        qcs = None
        if working_legend_name:
            wl = LCCS3Legend.objects.filter(name = working_legend_name).first()
            if not wl:
                print("retrieving from existdb instead via bfis-service calls")
                # BFIS_SERVICE_PORT = '8080'
                legends_get_url = 'http://{1}:{2}/bfis-service/rest/legends/{0}'.format(working_legend_name, request.get_host(), BFIS_SERVICE_PORT)
                get_resp = requests.get(legends_get_url)
                if get_resp.status_code == 200:
                    wl = LCCS3Legend()
                    xml_text = get_resp.json()['results']
                    wl.xml_text = xml_text
        else:
            wl = LCCS3Legend.objects.filter(active = True).first()
        if lcml_class_name:
            qcm = LCCS3Class.objects.filter(name = lcml_class_name).first()
            dd = perform_assessment(wl, qcm, advanced_options)
        else:
            if lcml_query_legend_name:
                qcm = LCCS3Class()
                ql = LCCS3Legend.objects.filter(name = lcml_query_legend_name).first()
                qcm.xml_text = ql.xml_text
                pprint.pprint(ql)
            else:
                if lcml_snippet:
                    lcml_snippet = lcml_snippet.replace("'",'"')
                    xml_fragment_text = lcml_snippet
                    if lcml_snippet.find('LC_Legend')==-1:
                        xml_fragment_text = '{0}{1}{2}'.format(xml_legend_template_part1, lcml_snippet, xml_legend_template_part2)
                    try:
                        xml_root = etree.fromstring(xml_fragment_text)
                    except Exception as e:
                        dd = {'error':'invalid_snippet'}
                    qcm = LCCS3Class()
                    qcm.xml_text = xml_fragment_text
                    pprint.pprint(xml_fragment_text)
                    dd = perform_assessment(wl, qcm, advanced_options)
        dd = perform_assessment(wl, qcm, advanced_options)
        invdd = invert_payload_order(dd)
        pprint.pprint(invdd)
        pprint.pprint("with params:")
        pprint.pprint(params)
    return JsonResponse(invdd)


@csrf_exempt
def class_information_extensiveness(request):
    dd = {}
    if request.method == 'GET':
        params = json.loads(request.body)
        pprint.pprint(params)
        
    return JsonResponse(invdd)

    
